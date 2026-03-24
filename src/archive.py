"""
Archive module.
Categorises headlines using Claude API and appends them to a growing Excel database.
"""

import anthropic
import logging
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

ARCHIVE_FILE = "headlines_archive.xlsx"

CATEGORIZE_PROMPT = """You are a news categorisation assistant. Given a list of Indonesian news headlines, assign each one a topic category.

Choose the most appropriate single category for each headline. Use short, consistent category names in English such as:
Politics, Economy, Energy, Defence/Security, Foreign Affairs, Legal/Judiciary, Transportation, Environment, Health, Education, Religion, Technology, Sports, Entertainment, Business, Disaster/Emergency, Social Affairs, or other categories as needed.

Headlines to categorise:
{headlines_json}

Respond with ONLY a valid JSON array, where each item has:
- "index": the headline number (starting from 0)
- "category": the assigned topic category

Example response:
[{{"index": 0, "category": "Politics"}}, {{"index": 1, "category": "Economy"}}]

Respond with ONLY the JSON array. No preamble, no markdown backticks, no explanation."""


def categorize_headlines(headlines_list: list[dict], api_key: str, model: str = "claude-sonnet-4-20250514") -> list[str]:
    """
    Send headlines to Claude API for topic categorisation.

    Args:
        headlines_list: List of dicts with 'title' and 'source' keys
        api_key: Anthropic API key
        model: Claude model to use

    Returns:
        List of category strings, one per headline
    """
    if not headlines_list:
        return []

    # Prepare headlines for the prompt
    headlines_for_prompt = [
        {"index": i, "title": h["title"], "source": h["source"]}
        for i, h in enumerate(headlines_list)
    ]

    try:
        client = anthropic.Anthropic(api_key=api_key)

        message = client.messages.create(
            model=model,
            max_tokens=2000,
            messages=[
                {
                    "role": "user",
                    "content": CATEGORIZE_PROMPT.format(
                        headlines_json=json.dumps(headlines_for_prompt, ensure_ascii=False)
                    )
                }
            ]
        )

        response_text = message.content[0].text.strip()

        # Clean up response in case Claude wraps it in markdown
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1]
            response_text = response_text.rsplit("```", 1)[0].strip()

        categories_data = json.loads(response_text)

        # Build category list matching headline order
        categories = ["Uncategorised"] * len(headlines_list)
        for item in categories_data:
            idx = item.get("index", -1)
            cat = item.get("category", "Uncategorised")
            if 0 <= idx < len(categories):
                categories[idx] = cat

        logger.info(f"Categorised {len(headlines_list)} headlines into topics")
        return categories

    except Exception as e:
        logger.error(f"Categorisation failed: {e}")
        return ["Uncategorised"] * len(headlines_list)


def save_to_excel(all_headlines: dict, categories: list[str], today_date: str, filepath: str = ARCHIVE_FILE):
    """
    Append today's headlines with categories to the Excel archive.

    Creates the file if it doesn't exist. Appends rows if it does.

    Columns: Date | Topic | Headline | Link
    """
    # Flatten headlines into a list with source info
    flat_headlines = []
    for source, headlines in all_headlines.items():
        for h in headlines:
            flat_headlines.append({
                "title": h.title,
                "source": source,
                "url": h.url,
            })

    # Load existing workbook or create new one
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Headlines Archive"

        # Style the header row
        headers = ["Date", "Topic", "Headline", "Link"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions["A"].width = 22  # Date
        ws.column_dimensions["B"].width = 22  # Topic
        ws.column_dimensions["C"].width = 80  # Headline
        ws.column_dimensions["D"].width = 50  # Link

        # Freeze header row
        ws.freeze_panes = "A2"

    # Append rows
    link_font = Font(color="1A73E8", underline="single")
    for i, headline in enumerate(flat_headlines):
        category = categories[i] if i < len(categories) else "Uncategorised"
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=today_date)
        ws.cell(row=row_num, column=2, value=category)
        ws.cell(row=row_num, column=3, value=headline["title"])

        # Create clickable hyperlink in the Link column
        link_cell = ws.cell(row=row_num, column=4, value=headline["url"])
        link_cell.hyperlink = headline["url"]
        link_cell.font = link_font

    # Auto-filter on all columns
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    wb.save(filepath)
    logger.info(f"Saved {len(flat_headlines)} headlines to {filepath} (total rows: {ws.max_row - 1})")


def archive_headlines(all_headlines: dict, api_key: str, today_date: str, model: str = "claude-sonnet-4-20250514", filepath: str = ARCHIVE_FILE):
    """
    Full archive pipeline: flatten headlines → categorise via AI → save to Excel.
    """
    # Flatten for categorisation
    flat_headlines = []
    for source, headlines in all_headlines.items():
        for h in headlines:
            flat_headlines.append({
                "title": h.title,
                "source": source,
            })

    # Get AI categories
    logger.info(f"Categorising {len(flat_headlines)} headlines...")
    categories = categorize_headlines(flat_headlines, api_key, model=model)

    # Save to Excel
    save_to_excel(all_headlines, categories, today_date, filepath=filepath)

    return categories
